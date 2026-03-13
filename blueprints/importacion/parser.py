"""
Parser de ficheros Excel para importación masiva de datos GMAO.
Cada función devuelve un dict keyed por nombre de hoja con listas de filas.
Cada fila es un dict con _fila (número de fila Excel) y los campos del encabezado.
"""
import io
import openpyxl


# Patrones de texto a ignorar en filas de instrucciones
SKIP_PATTERNS = [
    '\u2699',          # ⚙
    'GMAO \u00b7',     # GMAO ·
    'INSTRUCCIONES',
    'INSTRUCCI\u00d3N',  # INSTRUCCIÓN
    'Nivel ',
    'Una fila',
    'Lea estas',
    'Este fichero',
    'Campos OBLIGATORIOS',
    'Notas de importaci\u00f3n',  # Notas de importación
    'EJEMPLO',
]


def _should_skip_row(first_cell_value):
    """Devuelve True si la primera celda parece una fila de instrucciones."""
    first = str(first_cell_value or '').strip()
    if not first:
        return True
    for p in SKIP_PATTERNS:
        if first.startswith(p) or p.upper() in first.upper():
            return True
    return False


def _find_header_row(ws):
    """Encuentra la primera fila que parece un encabezado real (no instrucciones, no vacía)."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        first = str(row[0] or '').strip()
        if not first:
            continue
        skip = any(
            first.startswith(p) or p.upper() in first.upper()
            for p in SKIP_PATTERNS
        )
        if not skip:
            return row_idx, list(row)
    return None, None


def _normalize_header(val):
    """Convierte un encabezado a clave normalizada (minúsculas, sin espacios, sin acentos)."""
    if val is None:
        return None
    s = str(val).strip().lower()
    # Eliminar acentos básicos
    replacements = {
        '\xe1': 'a', '\xe9': 'e', '\xed': 'i', '\xf3': 'o', '\xfa': 'u',
        '\xc1': 'a', '\xc9': 'e', '\xcd': 'i', '\xd3': 'o', '\xda': 'u',
        '\xfc': 'u', '\xdc': 'u',
        '\xf1': 'n', '\xd1': 'n',
        ' ': '_', '/': '_', '-': '_', '(': '', ')': '', '.': '',
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s


def _parse_sheet(ws):
    """
    Parsea una hoja de cálculo y devuelve lista de dicts.
    Ignora filas de instrucciones y filas completamente vacías.
    Omite filas donde notas_importacion contiene 'OBLIGATORIO'.
    """
    header_row_idx, header = _find_header_row(ws)
    if header_row_idx is None:
        return []

    # Normalizar claves del encabezado
    keys = [_normalize_header(h) for h in header]

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= header_row_idx:
            continue

        # Comprobar si la fila está completamente vacía
        values = list(row)
        if all(v is None or str(v).strip() == '' for v in values):
            continue

        # Construir dict de la fila
        row_dict = {}
        for i, key in enumerate(keys):
            if key is None:
                continue
            val = values[i] if i < len(values) else None
            row_dict[key] = val

        # Omitir filas donde notas_importacion contenga 'OBLIGATORIO'
        notas = str(row_dict.get('notas_importacion', '') or '').strip().upper()
        if 'OBLIGATORIO' in notas:
            continue

        # Omitir si la primera celda significativa parece instrucción
        first_key = keys[0] if keys else None
        if first_key and _should_skip_row(row_dict.get(first_key)):
            continue

        row_dict['_fila'] = row_idx
        rows.append(row_dict)

    return rows


def _parse_workbook(file_bytes, expected_sheets=None):
    """
    Abre un workbook desde bytes y parsea sus hojas.
    Devuelve dict {nombre_hoja: [filas]}.
    Si expected_sheets es una lista, solo parsea esas hojas (insensible a mayúsculas).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        if expected_sheets:
            match = next(
                (e for e in expected_sheets if e.upper() == sheet_name.upper()),
                None
            )
            if match is None:
                continue
            canonical = match
        else:
            canonical = sheet_name

        ws = wb[sheet_name]
        result[canonical] = _parse_sheet(ws)

    return result


# =============================================================================
# FUNCIONES PÚBLICAS POR TIPO DE IMPORTACIÓN
# =============================================================================

def parse_activos(file_bytes):
    """
    Parsea fichero de Activos.
    Hojas esperadas: PLANTAS, ZONAS, LINEAS, MAQUINAS, ELEMENTOS
    """
    return _parse_workbook(file_bytes, ['PLANTAS', 'ZONAS', 'LINEAS', 'MAQUINAS', 'ELEMENTOS'])


def parse_gamas(file_bytes):
    """
    Parsea fichero de Gamas de Mantenimiento.
    Hojas esperadas: GAMAS, TAREAS, CHECKLIST, RECAMBIOS
    """
    return _parse_workbook(file_bytes, ['GAMAS', 'TAREAS', 'CHECKLIST', 'RECAMBIOS'])


def parse_historico(file_bytes):
    """
    Parsea fichero de Histórico de OTs.
    Hoja esperada: ORDENES
    """
    return _parse_workbook(file_bytes, ['ORDENES'])


def parse_recambios(file_bytes):
    """
    Parsea fichero de Recambios.
    Hoja esperada: RECAMBIOS
    """
    return _parse_workbook(file_bytes, ['RECAMBIOS'])


def parse_tecnicos(file_bytes):
    """
    Parsea fichero de Técnicos.
    Hoja esperada: TECNICOS
    """
    return _parse_workbook(file_bytes, ['TECNICOS'])


def parse_usuarios(file_bytes):
    """
    Parsea fichero de Usuarios.
    Hoja esperada: USUARIOS
    """
    return _parse_workbook(file_bytes, ['USUARIOS'])
