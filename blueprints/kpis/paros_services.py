"""
Cálculos de KPIs de paros de producción para el GMAO.

Fuente: OrdenTrabajo (tipo='correctivo', tiempoParada > 0)
Un paro de producción = OT correctiva con tiempoParada reportado al cierre.
El conteo es a nivel de OT (1 OT = 1 paro).
"""
import logging
import math
from calendar import monthrange
from datetime import date, datetime, timedelta

from sqlalchemy import and_, or_

from models import db, OrdenTrabajo, Maquina, Elemento, Linea, Zona, Planta, ConfiguracionGeneral

log = logging.getLogger(__name__)


# =============================================================================
# MAPA DE TURNOS (igual que en dashboard_services para consistencia)
# clave: 'horas/días' → (horas_dia, dias_semana)
# =============================================================================

_TURNOS = {
    '8/5':  (8,  5),
    '8/6':  (8,  6),
    '10/5': (10, 5),
    '12/5': (12, 5),
    '16/5': (16, 5),
    '16/6': (16, 6),
    '12/7': (12, 7),
    '24/5': (24, 5),
    '24/6': (24, 6),
    '24/7': (24, 7),
}


# =============================================================================
# CONSTANTES — REFERENCIAS CLASE MUNDIAL (no hardcodeadas en el template)
# =============================================================================

BENCHMARKS = {
    # clave: {ref, op ('gt'|'lt'), label, unidad}
    'mtbf_h':     {'ref': 100,   'op': 'gt', 'label': '> 100 h',      'unidad': 'h'},
    'mttr_h':     {'ref': 1,     'op': 'lt', 'label': '< 1 h',        'unidad': 'h'},
    'disp_pct':   {'ref': 95,    'op': 'gt', 'label': '> 95 %',       'unidad': '%'},
    'r_24h':      {'ref': 80,    'op': 'gt', 'label': '> 80 %',       'unidad': '%'},
    'r_168h':     {'ref': 30,    'op': 'gt', 'label': '> 30 %',       'unidad': '%'},
    'paros_dia':  {'ref': 0.5,   'op': 'lt', 'label': '< 0.5 /día',   'unidad': '/día'},
    'hparos_dia': {'ref': 0.5,   'op': 'lt', 'label': '< 0.5 h/día',  'unidad': 'h/día'},
    'lambda':     {'ref': 0.010, 'op': 'lt', 'label': '< 0.010',      'unidad': 'f/h'},
}

# Orden y etiquetas para la tabla benchmarking
BENCH_ROWS = [
    ('MTBF',                    'mtbf_h'),
    ('MTTR',                    'mttr_h'),
    ('Disponibilidad',          'disp_pct'),
    ('Fiabilidad 24 h',         'r_24h'),
    ('Fiabilidad semanal 168 h','r_168h'),
    ('Paros / día',             'paros_dia'),
    ('Horas paro / día',        'hparos_dia'),
    ('Tasa de fallos (λ)',      'lambda'),
]


# =============================================================================
# HELPERS DE FECHAS Y PERIODOS
# =============================================================================

def _parse_fecha(s):
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except Exception:
        return None


def _fecha_referencia(ot):
    """Fecha efectiva de la OT: fechaFin si existe, si no fechaCreacion."""
    d = ot.fechaFin or ot.fechaCreacion
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return date.today()


def _horas_operativas(fecha_ini: date, fecha_fin: date, turno_key: str) -> float:
    """
    Horas operativas reales entre dos fechas según el turno configurado en
    ConfiguracionGeneral (clave 'turno_planta').

    Si dias_semana < 7 solo cuenta los días laborables:
      /5 → lun-vie, /6 → lun-sáb, /7 → todos los días.
    """
    horas_dia, dias_semana = _TURNOS.get(turno_key, (24, 7))
    total = 0.0
    current = fecha_ini
    while current <= fecha_fin:
        dow = current.weekday()  # 0=lun … 6=dom
        if dias_semana == 7:
            total += horas_dia
        elif dias_semana == 6 and dow < 6:   # lun-sáb
            total += horas_dia
        elif dias_semana == 5 and dow < 5:   # lun-vie
            total += horas_dia
        current += timedelta(days=1)
    return total


def _dias_periodo(fecha_ini: date, fecha_fin: date) -> int:
    return (fecha_fin - fecha_ini).days + 1


def _periodos_mensuales(fecha_ini: date, fecha_fin: date):
    """Genera lista de (key, label, date_ini, date_fin) por mes."""
    result = []
    cur = fecha_ini.replace(day=1)
    while cur <= fecha_fin:
        days_in_month = monthrange(cur.year, cur.month)[1]
        p_ini = max(cur, fecha_ini)
        p_fin = min(date(cur.year, cur.month, days_in_month), fecha_fin)
        key   = cur.strftime('%Y-%m')
        label = cur.strftime('%b %Y')
        result.append((key, label, p_ini, p_fin))
        cur = date(cur.year + 1, 1, 1) if cur.month == 12 else date(cur.year, cur.month + 1, 1)
    return result


def _periodos_anuales(fecha_ini: date, fecha_fin: date):
    """Genera lista de (key, label, date_ini, date_fin) por año."""
    result = []
    for year in range(fecha_ini.year, fecha_fin.year + 1):
        p_ini = max(date(year, 1, 1), fecha_ini)
        p_fin = min(date(year, 12, 31), fecha_fin)
        result.append((str(year), str(year), p_ini, p_fin))
    return result


# =============================================================================
# CÁLCULOS KPI POR PERIODO
# =============================================================================

def _calcular_kpis_periodo(n_paros: int, h_paros: float, h_cal: float, dias: int) -> dict:
    """
    Devuelve dict con todos los KPIs de un periodo dado.
    Si n_paros == 0 devuelve "—" para los indicadores que requieren datos.
    """
    if h_paros > h_cal and h_cal > 0:
        log.warning(
            "ANOMALIA: h_paros=%.2f > h_cal=%.2f (dias=%d)", h_paros, h_cal, dias
        )

    if n_paros == 0:
        return {
            'n_paros': 0, 'h_paros': 0.0,
            'h_cal': round(h_cal, 2), 'h_func': round(h_cal, 2),
            'mtbf_h': None, 'mtbf_dias': None, 'mttr_h': None,
            'disp_pct': 100.0,
            'lambda': 0.0, 'r_24h': 100.0, 'r_168h': 100.0,
            'paros_dia': 0.0, 'hparos_dia': 0.0,
        }

    h_func  = h_cal - h_paros
    mtbf_h  = h_func / n_paros
    mttr_h  = h_paros / n_paros
    disp    = mtbf_h / (mtbf_h + mttr_h) * 100.0
    lam     = n_paros / h_cal if h_cal > 0 else 0.0
    r_24h   = math.exp(-lam * 24)   * 100.0
    r_168h  = math.exp(-lam * 168)  * 100.0

    # Validación cruzada interna
    disp_cross = (h_func / h_cal * 100.0) if h_cal > 0 else 0.0
    if abs(disp - disp_cross) > 0.1:
        log.warning(
            "INCONSISTENCIA disponibilidad: MTBF/(MTBF+MTTR)=%.4f%% vs H_func/H_cal=%.4f%%",
            disp, disp_cross
        )

    return {
        'n_paros':    n_paros,
        'h_paros':    round(h_paros, 2),
        'h_cal':      round(h_cal, 2),
        'h_func':     round(h_func, 2),
        'mtbf_h':     round(mtbf_h, 2),
        'mtbf_dias':  round(mtbf_h / 24.0, 2),
        'mttr_h':     round(mttr_h, 2),
        'disp_pct':   round(disp, 2),
        'lambda':     round(lam, 6),
        'r_24h':      round(r_24h, 2),
        'r_168h':     round(r_168h, 2),
        'paros_dia':  round(n_paros / dias, 4) if dias > 0 else 0.0,
        'hparos_dia': round(h_paros / dias, 4) if dias > 0 else 0.0,
    }


def _add_deltas(periodos_data: list) -> list:
    """Añade campos delta_* comparando cada periodo con el anterior."""
    for i, p in enumerate(periodos_data):
        if i == 0:
            p.update(delta_n_pct=None, delta_h_pct=None,
                     delta_mtbf_pct=None, delta_disp_pp=None)
            continue
        prev = periodos_data[i - 1]

        def _pct(curr, ant):
            if ant is None or curr is None or ant == 0:
                return None
            return round((curr - ant) / abs(ant) * 100.0, 1)

        def _pp(curr, ant):
            if ant is None or curr is None:
                return None
            return round(curr - ant, 2)

        p['delta_n_pct']    = _pct(p['n_paros'],   prev['n_paros'])
        p['delta_h_pct']    = _pct(p['h_paros'],   prev['h_paros'])
        p['delta_mtbf_pct'] = _pct(p.get('mtbf_h'), prev.get('mtbf_h'))
        p['delta_disp_pp']  = _pp(p.get('disp_pct'), prev.get('disp_pct'))
    return periodos_data


def _tendencia(valores: list, inverso: bool = False) -> str:
    """
    Tendencia entre último y penúltimo valor no nulo.
    inverso=True: para MTTR y λ (bajar es mejorar → devuelve 'mejora').
    Retorna: 'mejora' | 'estable' | 'deterioro'
    """
    vals = [v for v in valores if v is not None]
    if len(vals) < 2:
        return 'estable'
    ultimo, penultimo = vals[-1], vals[-2]
    if penultimo == 0:
        return 'estable'
    delta_pct = (ultimo - penultimo) / abs(penultimo) * 100.0
    if inverso:
        delta_pct = -delta_pct
    if delta_pct > 5.0:
        return 'mejora'
    if delta_pct < -5.0:
        return 'deterioro'
    return 'estable'


# =============================================================================
# BENCHMARKING
# =============================================================================

def _estado_benchmark(valor, ref, op: str) -> str:
    """Retorna 'ok' | 'mejorable' | 'critico'."""
    if valor is None:
        return 'nd'
    if op == 'gt':
        if valor >= ref:
            return 'ok'
        if valor >= ref * 0.90:
            return 'mejorable'
        return 'critico'
    else:  # lt
        if valor <= ref:
            return 'ok'
        if valor <= ref * 1.10:
            return 'mejorable'
        return 'critico'


def _gap_fmt(valor, ref, op: str, unidad: str) -> str:
    if valor is None:
        return '—'
    gap = (valor - ref) if op == 'gt' else (ref - valor)
    sign = '+' if gap >= 0 else ''
    return f"{sign}{round(gap, 3)} {unidad}"


def _calcular_benchmarking(kpis: dict) -> dict:
    """Genera la tabla de benchmarking comparando kpis con referencias clase mundial."""
    rows = []
    for nombre, campo in BENCH_ROWS:
        bm = BENCHMARKS.get(campo)
        if not bm:
            continue
        valor  = kpis.get(campo)
        ref    = bm['ref']
        op     = bm['op']
        estado = _estado_benchmark(valor, ref, op)
        rows.append({
            'indicador':  nombre,
            'campo':      campo,
            'valor':      valor,
            'valor_fmt':  f"{valor:.3f}" if valor is not None else '—',
            'referencia': bm['label'],
            'unidad':     bm['unidad'],
            'gap':        _gap_fmt(valor, ref, op, bm['unidad']),
            'estado':     estado,
        })
    return {'rows': rows, 'kpis_globales': kpis}


# =============================================================================
# RESOLUCIÓN DE JERARQUÍA
# =============================================================================

def _precargar_jerarquia():
    """Carga en memoria los modelos de jerarquía para evitar N+1 queries."""
    maquinas = {m.id: m for m in Maquina.query.all()}
    elementos = {e.id: e for e in Elemento.query.all()}
    lineas   = {l.id: l for l in Linea.query.all()}
    return maquinas, elementos, lineas


def _resolver_linea_maquina(ot, maquinas: dict, elementos: dict):
    """
    Devuelve (linea_id, maquina_id, maquina_nombre) para una OT.
    Sube la jerarquía desde el equipoTipo/equipoId.
    """
    et  = ot.equipoTipo
    eid = ot.equipoId

    if et == 'maquina':
        maq = maquinas.get(eid)
        if maq:
            return maq.lineaId, maq.id, maq.nombre

    elif et == 'elemento':
        elem = elementos.get(eid)
        if elem:
            maq = maquinas.get(elem.maquinaId)
            if maq:
                return maq.lineaId, maq.id, maq.nombre

    elif et == 'linea':
        return eid, None, ''

    # Fallback al campo legacy maquinaId
    if ot.maquinaId:
        maq = maquinas.get(ot.maquinaId)
        if maq:
            return maq.lineaId, maq.id, maq.nombre

    return None, None, ''


# =============================================================================
# CONSULTA PRINCIPAL
# =============================================================================

def _query_paros(fecha_ini: date, fecha_fin: date):
    """
    Devuelve todas las OTs que son paros de producción en el rango de fechas dado.
    Filtro fijo: tipo='correctivo' AND tiempoParada > 0.
    Fecha de referencia: fechaFin si existe, si no fechaCreacion.
    """
    fi_dt = datetime.combine(fecha_ini, datetime.min.time())
    ff_dt = datetime.combine(fecha_fin, datetime.max.time())

    date_filter = or_(
        and_(
            OrdenTrabajo.fechaFin.isnot(None),
            OrdenTrabajo.fechaFin >= fi_dt,
            OrdenTrabajo.fechaFin <= ff_dt,
        ),
        and_(
            OrdenTrabajo.fechaFin.is_(None),
            OrdenTrabajo.fechaCreacion >= fi_dt,
            OrdenTrabajo.fechaCreacion <= ff_dt,
        ),
    )

    return OrdenTrabajo.query.filter(
        OrdenTrabajo.tipo == 'correctivo',
        OrdenTrabajo.tiempoParada > 0,
        date_filter,
    ).all()


# =============================================================================
# PUNTO DE ENTRADA PRINCIPAL
# =============================================================================

def calcular_paros(
    fecha_ini: date,
    fecha_fin: date,
    agrupacion: str = 'mensual',
    lineas_ids: list = None,
    maquinas_ids: list = None,
    plantas_ids: list = None,
    zonas_ids: list = None,
) -> dict:
    """
    Calcula todos los KPIs de paros de producción.

    El régimen de turnos (horas calendario) se lee automáticamente de
    ConfiguracionGeneral (clave 'turno_planta').

    Parámetros
    ----------
    fecha_ini, fecha_fin : rango de fechas
    agrupacion           : 'mensual' | 'anual'
    plantas_ids          : lista de int — filtro por planta (opcional)
    zonas_ids            : lista de int — filtro por zona (opcional)
    lineas_ids           : lista de int — filtro por línea (opcional)
    maquinas_ids         : lista de int — filtro por máquina (opcional)

    Retorna
    -------
    dict con claves: periodos_labels, periodos_keys, global, por_grupo, top10, benchmarking
    """
    turno_key = ConfiguracionGeneral.obtener('turno_planta', '24/7')
    # 1. Cargar OTs base
    ots_all = _query_paros(fecha_ini, fecha_fin)

    # 2. Cargar jerarquía una vez
    maquinas, elementos, lineas_dict = _precargar_jerarquia()
    zonas_dict  = {z.id: z for z in Zona.query.all()}
    plantas_dict = {p.id: p for p in Planta.query.all()}

    # 3. Construir sets de filtro — resolver cadena Planta→Zona→Línea
    # Si se filtra por planta/zona, expandimos a lineas_ids automáticamente
    lineas_set   = set(lineas_ids)   if lineas_ids   else None
    maquinas_set = set(maquinas_ids) if maquinas_ids else None
    plantas_set  = set(plantas_ids)  if plantas_ids  else None
    zonas_set    = set(zonas_ids)    if zonas_ids    else None

    # Si hay filtro de Planta o Zona, calculamos el conjunto de lineas válidas
    if plantas_set or zonas_set:
        lineas_validas = set()
        for lid, linea in lineas_dict.items():
            zona = zonas_dict.get(linea.zonaId)
            if zona is None:
                continue
            if zonas_set and zona.id not in zonas_set:
                continue
            if plantas_set and zona.plantaId not in plantas_set:
                continue
            lineas_validas.add(lid)
        # Intersectar con filtro de linea explícito si existe
        lineas_set = (lineas_set & lineas_validas) if lineas_set else lineas_validas

    ots = []
    for ot in ots_all:
        lid, mid, _ = _resolver_linea_maquina(ot, maquinas, elementos)
        if lineas_set   and lid not in lineas_set:
            continue
        if maquinas_set and mid not in maquinas_set:
            continue
        ots.append(ot)

    # 4. Generar periodos
    periodos = (
        _periodos_anuales(fecha_ini, fecha_fin)
        if agrupacion == 'anual'
        else _periodos_mensuales(fecha_ini, fecha_fin)
    )

    # 5. Inicializar acumuladores
    periodo_global  = {key: {'n_paros': 0, 'h_paros': 0.0} for key, *_ in periodos}
    periodo_linea   = {key: {} for key, *_ in periodos}
    periodo_maquina = {key: {} for key, *_ in periodos}
    top_acum        = {}   # (lid, mid) → acumulado total del periodo completo

    def _acum(d: dict, k, h: float):
        if k not in d:
            d[k] = {'n_paros': 0, 'h_paros': 0.0}
        d[k]['n_paros'] += 1
        d[k]['h_paros'] += h

    # 6. Distribuir cada OT en su periodo
    for ot in ots:
        ref_date = _fecha_referencia(ot)
        p_key = None
        for key, _label, p_ini, p_fin in periodos:
            if p_ini <= ref_date <= p_fin:
                p_key = key
                break
        if p_key is None:
            continue

        h   = float(ot.tiempoParada or 0.0)
        lid, mid, maq_nombre = _resolver_linea_maquina(ot, maquinas, elementos)

        # Global
        pg = periodo_global[p_key]
        pg['n_paros'] += 1
        pg['h_paros'] += h

        # Por línea
        if lid:
            _acum(periodo_linea[p_key], lid, h)

        # Por máquina
        if mid:
            _acum(periodo_maquina[p_key], mid, h)

        # Top acumulado (nivel de OT, periodo completo)
        tk = (lid, mid)
        if tk not in top_acum:
            linea_nom = lineas_dict[lid].nombre if lid and lid in lineas_dict else 'Sin línea'
            top_acum[tk] = {
                'linea_id':   lid,
                'maquina_id': mid,
                'linea':      linea_nom,
                'maquina':    maq_nombre,
                'n_paros':    0,
                'h_paros':    0.0,
            }
        top_acum[tk]['n_paros'] += 1
        top_acum[tk]['h_paros'] += h

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 1 — Indicadores globales por periodo
    # ─────────────────────────────────────────────────────────────────────────
    global_data = []
    for key, label, p_ini, p_fin in periodos:
        pg   = periodo_global[key]
        h_cal = _horas_operativas(p_ini, p_fin, turno_key)
        dias  = _dias_periodo(p_ini, p_fin)
        kpis  = _calcular_kpis_periodo(pg['n_paros'], pg['h_paros'], h_cal, dias)
        kpis['key']   = key
        kpis['label'] = label
        global_data.append(kpis)

    _add_deltas(global_data)

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 2 — Por línea o por máquina
    # ─────────────────────────────────────────────────────────────────────────
    agrupado_por = 'maquina' if lineas_set else 'linea'

    if agrupado_por == 'linea':
        todas_ids = set()
        for key, *_ in periodos:
            todas_ids.update(periodo_linea[key].keys())
        acc_src = periodo_linea
        def _nombre(id_):
            l = lineas_dict.get(id_)
            return l.nombre if l else f'Línea {id_}'
    else:
        todas_ids = set()
        for key, *_ in periodos:
            todas_ids.update(periodo_maquina[key].keys())
        acc_src = periodo_maquina
        def _nombre(id_):
            m = maquinas.get(id_)
            return m.nombre if m else f'Máq. {id_}'

    grupos = []
    for gid in sorted(todas_ids):
        periodos_grupo = []
        for key, label, p_ini, p_fin in periodos:
            pg    = acc_src[key].get(gid, {'n_paros': 0, 'h_paros': 0.0})
            h_cal = _horas_operativas(p_ini, p_fin, turno_key)
            dias  = _dias_periodo(p_ini, p_fin)
            kpis  = _calcular_kpis_periodo(pg['n_paros'], pg['h_paros'], h_cal, dias)
            kpis['key']   = key
            kpis['label'] = label
            periodos_grupo.append(kpis)

        grupos.append({
            'id':               gid,
            'nombre':           _nombre(gid),
            'periodos':         periodos_grupo,
            'tendencia_mtbf':   _tendencia([p.get('mtbf_h')  for p in periodos_grupo], inverso=False),
            'tendencia_mttr':   _tendencia([p.get('mttr_h')  for p in periodos_grupo], inverso=True),
            'tendencia_disp':   _tendencia([p.get('disp_pct') for p in periodos_grupo], inverso=False),
            'tendencia_lambda': _tendencia([p.get('lambda')  for p in periodos_grupo], inverso=True),
        })

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 3 — Top 10 equipos
    # ─────────────────────────────────────────────────────────────────────────
    total_paros = sum(v['n_paros'] for v in top_acum.values())
    top10 = sorted(top_acum.values(), key=lambda x: x['n_paros'], reverse=True)[:10]

    acc_pct = 0.0
    for t in top10:
        t['pct_total']     = round(t['n_paros'] / total_paros * 100.0, 1) if total_paros else 0.0
        t['h_paros']       = round(t['h_paros'], 2)
        acc_pct           += t['pct_total']
        t['pct_acumulado'] = round(acc_pct, 1)

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 4 — Benchmarking global del periodo completo
    # ─────────────────────────────────────────────────────────────────────────
    total_n   = sum(p['n_paros']  for p in global_data)
    total_h   = sum(p['h_paros']  for p in global_data)
    total_cal = sum(p['h_cal']    for p in global_data)
    total_dias = _dias_periodo(fecha_ini, fecha_fin)
    bench_kpis = _calcular_kpis_periodo(total_n, total_h, total_cal, total_dias)
    benchmarking = _calcular_benchmarking(bench_kpis)

    return {
        'periodos_labels': [p[1] for p in periodos],
        'periodos_keys':   [p[0] for p in periodos],
        'global':          global_data,
        'por_grupo': {
            'agrupado_por': agrupado_por,
            'grupos':       grupos,
        },
        'top10':         top10,
        'benchmarking':  benchmarking,
        'filtros_aplicados': {
            'fecha_ini':    fecha_ini.isoformat(),
            'fecha_fin':    fecha_fin.isoformat(),
            'agrupacion':   agrupacion,
            'turno_planta': turno_key,
            'plantas_ids':  list(plantas_set)   if plantas_set  else [],
            'zonas_ids':    list(zonas_set)     if zonas_set    else [],
            'lineas_ids':   list(lineas_set)    if lineas_set   else [],
            'maquinas_ids': list(maquinas_set)  if maquinas_set else [],
        },
    }


# =============================================================================
# LISTAS PARA LOS FILTROS
# =============================================================================

def get_plantas():
    """Lista {id, nombre} de todas las plantas."""
    return [
        {'id': p.id, 'nombre': p.nombre}
        for p in Planta.query.order_by(Planta.nombre).all()
    ]


def get_zonas(plantas_ids: list = None):
    """Lista {id, nombre, planta_id} de zonas, filtradas por planta si se indica."""
    q = Zona.query.order_by(Zona.nombre)
    if plantas_ids:
        q = q.filter(Zona.plantaId.in_(plantas_ids))
    return [{'id': z.id, 'nombre': z.nombre, 'planta_id': z.plantaId} for z in q.all()]


def get_lineas(zonas_ids: list = None):
    """Lista {id, nombre, zona_id} de todas las líneas, filtradas por zona si se indica."""
    q = Linea.query.order_by(Linea.nombre)
    if zonas_ids:
        q = q.filter(Linea.zonaId.in_(zonas_ids))
    return [{'id': l.id, 'nombre': l.nombre, 'zona_id': l.zonaId} for l in q.all()]


def get_maquinas(lineas_ids: list = None):
    """Lista {id, nombre, linea_id} de máquinas, filtradas por línea si se indica."""
    q = Maquina.query.order_by(Maquina.nombre)
    if lineas_ids:
        q = q.filter(Maquina.lineaId.in_(lineas_ids))
    return [{'id': m.id, 'nombre': m.nombre, 'linea_id': m.lineaId} for m in q.all()]


# =============================================================================
# EXPORTACIÓN EXCEL
# =============================================================================

def exportar_paros_excel(datos: dict) -> 'io.BytesIO':
    """Genera un fichero Excel con las tablas del análisis de paros."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill('solid', fgColor='1565C0')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
    BOLD      = Font(bold=True, size=10)
    NORM      = Font(size=10)
    TOTAL_FILL = PatternFill('solid', fgColor='E3F2FD')

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    def _val(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = NORM
        c.alignment = Alignment(horizontal='right')

    periodos     = datos.get('periodos_labels', [])
    global_data  = datos.get('global', [])

    # ── Hoja 1: Indicadores Globales ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Global'

    FILAS_KPI = [
        ('Nº Paros',           'n_paros',    ''),
        ('Horas Paro',         'h_paros',    'h'),
        ('Paros/día',          'paros_dia',  '/día'),
        ('Horas paro/día',     'hparos_dia', 'h/día'),
        ('MTBF',               'mtbf_h',     'h'),
        ('MTBF',               'mtbf_dias',  'días'),
        ('MTTR',               'mttr_h',     'h'),
        ('Disponibilidad',     'disp_pct',   '%'),
        ('Tasa fallos (λ)',     'lambda',     'f/h'),
        ('Fiabilidad 24h',     'r_24h',      '%'),
        ('Fiabilidad 168h',    'r_168h',     '%'),
        ('Δ Nº Paros',         'delta_n_pct',    '%'),
        ('Δ Horas Paro',       'delta_h_pct',    '%'),
        ('Δ MTBF',             'delta_mtbf_pct', '%'),
        ('Δ Disponibilidad',   'delta_disp_pp',  'pp'),
    ]

    _hdr(ws1, 1, 1, 'Indicador')
    _hdr(ws1, 1, 2, 'Unidad')
    for j, lbl in enumerate(periodos, start=3):
        _hdr(ws1, 1, j, lbl)

    for i, (nombre, campo, unidad) in enumerate(FILAS_KPI, start=2):
        ws1.cell(row=i, column=1, value=nombre).font = BOLD
        ws1.cell(row=i, column=2, value=unidad).font = NORM
        for j, p in enumerate(global_data, start=3):
            v = p.get(campo)
            _val(ws1, i, j, v)

    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 8
    for j in range(3, 3 + len(periodos)):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 14

    # ── Hoja 2: Por Grupo ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Por Línea-Máquina')
    grupos     = datos.get('por_grupo', {}).get('grupos', [])
    agrup      = datos.get('por_grupo', {}).get('agrupado_por', 'linea')
    titulo_col = 'Línea' if agrup == 'linea' else 'Máquina'

    GRUPO_KPIS = [
        ('MTBF (h)',    'mtbf_h'),
        ('MTTR (h)',    'mttr_h'),
        ('Disp. (%)',   'disp_pct'),
        ('Nº Paros',    'n_paros'),
        ('Horas Paro',  'h_paros'),
        ('λ (f/h)',     'lambda'),
    ]

    row = 1
    for kpi_label, kpi_campo in GRUPO_KPIS:
        _hdr(ws2, row, 1, f'{kpi_label} por {titulo_col}')
        for j, lbl in enumerate(periodos, start=2):
            _hdr(ws2, row, j, lbl)
        _hdr(ws2, row, 2 + len(periodos), 'Tendencia')
        row += 1

        for g in grupos:
            ws2.cell(row=row, column=1, value=g['nombre']).font = BOLD
            for j, p in enumerate(g['periodos'], start=2):
                v = p.get(kpi_campo)
                _val(ws2, row, j, v)
            tend_key = f"tendencia_{kpi_campo.split('_')[0]}"
            tend = g.get(tend_key, '')
            ws2.cell(row=row, column=2 + len(periodos), value=tend).font = NORM
            row += 1
        row += 1  # fila vacía entre secciones

    ws2.column_dimensions['A'].width = 25
    for j in range(2, 2 + len(periodos) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 14

    # ── Hoja 3: Top 10 ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Top 10 Equipos')
    for j, hdr in enumerate(['Equipo (Línea - Máquina)', 'Nº Paros', '% del Total', 'Horas Paro'], start=1):
        _hdr(ws3, 1, j, hdr)
    for i, t in enumerate(datos.get('top10', []), start=2):
        equipo = f"{t.get('linea', '')} - {t.get('maquina', '')}" if t.get('maquina') else t.get('linea', '—')
        ws3.cell(row=i, column=1, value=equipo).font = NORM
        _val(ws3, i, 2, t.get('n_paros'))
        _val(ws3, i, 3, t.get('pct_total'))
        _val(ws3, i, 4, t.get('h_paros'))
    ws3.column_dimensions['A'].width = 35
    for j in range(2, 5):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 14

    # ── Hoja 4: Benchmarking ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('Benchmarking')
    for j, hdr in enumerate(['Indicador', 'Valor Actual', 'Unidad', 'Referencia Clase Mundial', 'Gap', 'Estado'], start=1):
        _hdr(ws4, 1, j, hdr)
    ESTADO_MAP = {'ok': 'OK', 'mejorable': 'Mejorable', 'critico': 'Critico', 'nd': 'N/D'}
    for i, row_data in enumerate(datos.get('benchmarking', {}).get('rows', []), start=2):
        ws4.cell(row=i, column=1, value=row_data.get('indicador')).font = BOLD
        _val(ws4, i, 2, row_data.get('valor'))
        ws4.cell(row=i, column=3, value=row_data.get('unidad')).font = NORM
        ws4.cell(row=i, column=4, value=row_data.get('referencia')).font = NORM
        ws4.cell(row=i, column=5, value=row_data.get('gap')).font = NORM
        ws4.cell(row=i, column=6, value=ESTADO_MAP.get(row_data.get('estado', ''), '')).font = NORM
    for j, w in enumerate([28, 14, 10, 26, 14, 12], start=1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =============================================================================
# EXPORTACIÓN PDF
# =============================================================================

def _png_size(data: bytes):
    """Devuelve (ancho, alto) en píxeles leyendo la cabecera PNG.
    Estructura: [0-7 firma] [8-11 lon] [12-15 'IHDR'] [16-19 width] [20-23 height]
    """
    import struct
    if len(data) >= 24 and data[:8] == b'\x89PNG\r\n\x1a\n':
        return struct.unpack('>II', data[16:24])   # width @ 16, height @ 20
    return (800, 400)


def exportar_paros_pdf(datos: dict, chart_images: dict = None) -> 'io.BytesIO':
    """
    Genera un PDF con las tablas de indicadores de paros.
    Cabecera con logo GMAO JGG en cada página.
    Incluye: Indicadores Globales, Por Línea/Máquina, Top 10 y Benchmarking.
    """
    import io
    import os
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, KeepTogether, PageBreak, Image as RLImage
    )

    # ── Constantes de color ───────────────────────────────────────────────────
    AZUL_HDR   = colors.HexColor('#00335F')   # color del logo (cabecera canvas)
    AZUL       = colors.HexColor('#1565C0')   # cabeceras de tabla
    AZUL_CLARO = colors.HexColor('#E3F2FD')
    GRIS_TXT   = colors.HexColor('#555555')
    BLANCO     = colors.white
    VERDE      = colors.HexColor('#2E7D32')
    NARANJA    = colors.HexColor('#E65100')
    ROJO       = colors.HexColor('#C62828')

    PAGE_W, PAGE_H = landscape(A4)
    HDR_H   = 1.6 * cm   # alto de la cabecera dibujada en canvas
    HDR_TOP = 0.35 * cm  # margen superior hasta la cabecera

    logo_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        'static', 'images', 'logoWhite.png'
    )

    filtros_ap = datos.get('filtros_aplicados', {})
    fi  = filtros_ap.get('fecha_ini', '—')
    ff  = filtros_ap.get('fecha_fin', '—')
    agr = filtros_ap.get('agrupacion', '—')
    tur = filtros_ap.get('turno_planta', '—')

    # ── Cabecera + pie dibujados en canvas (se repiten en cada página) ────────
    def _page_deco(canvas, doc):
        canvas.saveState()

        # Fondo de la cabecera con el color del logo
        hdr_y = PAGE_H - HDR_TOP - HDR_H
        canvas.setFillColor(AZUL_HDR)
        canvas.rect(1.2 * cm, hdr_y, PAGE_W - 2.4 * cm, HDR_H, fill=1, stroke=0)

        # Logo
        logo_drawn = False
        if os.path.exists(logo_path):
            try:
                canvas.drawImage(
                    logo_path,
                    1.6 * cm, hdr_y + 0.15 * cm,
                    width=2.8 * cm, height=HDR_H - 0.3 * cm,
                    preserveAspectRatio=True, mask='auto',
                )
                logo_drawn = True
            except Exception:
                pass

        # Texto "GMAO JGG"
        txt_x = (1.6 + 3.2) * cm if logo_drawn else 1.8 * cm
        canvas.setFillColor(BLANCO)
        canvas.setFont('Helvetica-Bold', 12)
        canvas.drawString(txt_x, hdr_y + HDR_H * 0.52, 'GMAO JGG')

        # Subtítulo
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(colors.HexColor('#8BB8D4'))
        canvas.drawString(txt_x, hdr_y + HDR_H * 0.16,
                          f'KPIs Paros de Producción  ·  {fi} → {ff}  ·  {agr}  ·  Turno: {tur}')

        # Pie de página
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(GRIS_TXT)
        canvas.drawString(1.5 * cm, 0.65 * cm,
                          f'GMAO JGG  ·  KPIs Paros de Producción  ·  {fi} → {ff}')
        canvas.drawRightString(PAGE_W - 1.5 * cm, 0.65 * cm, f'Pág. {doc.page}')

        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=HDR_TOP + HDR_H + 0.6 * cm,
        bottomMargin=1.5 * cm,
        title='KPIs Paros de Producción',
        author='GMAO JGG',
    )

    styles = getSampleStyleSheet()
    st_sec  = ParagraphStyle('sec', parent=styles['Heading2'],
                             fontSize=10, textColor=AZUL,
                             spaceBefore=8, spaceAfter=4)
    st_sub  = ParagraphStyle('sub', parent=styles['Normal'],
                             fontSize=8, textColor=GRIS_TXT,
                             spaceBefore=8, spaceAfter=3)

    periodos    = datos.get('periodos_labels', [])
    global_data = datos.get('global', [])

    import base64 as _b64

    chart_images = chart_images or {}
    log.info("PDF: %d gráficas recibidas: %s", len(chart_images), list(chart_images.keys()))

    # ── Helper: convierte base64 PNG → Image flowable con ancho fijo ──────────
    def _chart_img(key: str, target_w: float):
        b64 = chart_images.get(key, '')
        if not b64:
            return None
        try:
            raw = _b64.b64decode(b64.split(',', 1)[-1])
            pw, ph = _png_size(raw)
            target_h = target_w * ph / pw if pw else target_w * 0.5
            img = RLImage(io.BytesIO(raw), width=target_w, height=target_h)
            log.info("  gráfica '%s': %dx%d px → %.1fpt × %.1fpt", key, pw, ph, target_w, target_h)
            return img
        except Exception as e:
            log.warning("  ERROR gráfica '%s': %s", key, e)
            return None

    def _charts_row(keys: list, page_width: float):
        """Tabla de 1 fila con las gráficas indicadas en columnas iguales."""
        n = len(keys)
        gap = 0.3 * cm
        cell_w = (page_width - gap * (n - 1)) / n
        imgs = [_chart_img(k, cell_w) for k in keys]
        if not any(imgs):
            return None
        cells  = [img if img else '' for img in imgs]
        col_ws = [cell_w] * (n - 1) + [cell_w + gap * (n - 1)]
        tbl = Table([cells], colWidths=col_ws)
        tbl.setStyle(TableStyle([
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
        ]))
        return tbl

    # Estilo base para todas las tablas
    def _make_th_style():
        return [
            ('BACKGROUND',    (0, 0), (-1, 0), AZUL),
            ('TEXTCOLOR',     (0, 0), (-1, 0), BLANCO),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 7.5),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN',         (0, 1), (0, -1), 'LEFT'),
            ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [BLANCO, AZUL_CLARO]),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]

    story = []

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 1 — Indicadores Globales
    # Tabla con repetición de cabecera en cada página (splitByRow por defecto=1)
    # ─────────────────────────────────────────────────────────────────────────
    FILAS_GLOBAL_PDF = [
        ('Nº Paros',            'n_paros',    '{:.0f}'),
        ('Horas Paro',          'h_paros',    '{:.2f} h'),
        ('Paros / día',         'paros_dia',  '{:.3f}'),
        ('Horas paro / día',    'hparos_dia', '{:.3f} h/día'),
        ('MTBF (h)',            'mtbf_h',     '{:.2f} h'),
        ('MTBF (días)',         'mtbf_dias',  '{:.2f} días'),
        ('MTTR (h)',            'mttr_h',     '{:.2f} h'),
        ('Disponibilidad (%)',  'disp_pct',   '{:.2f} %'),
        ('Tasa fallos λ (f/h)', 'lambda',     '{:.5f}'),
        ('Fiabilidad 24 h (%)', 'r_24h',      '{:.2f} %'),
        ('Fiabilidad 168h (%)', 'r_168h',     '{:.2f} %'),
    ]

    col_ancho = (doc.width - 4.5 * cm) / max(len(periodos), 1)
    g1_rows = [['Indicador'] + periodos]
    for label, campo, fmt in FILAS_GLOBAL_PDF:
        fila = [label]
        for p in global_data:
            v = p.get(campo)
            fila.append(fmt.format(v) if v is not None else '—')
        g1_rows.append(fila)

    col_widths_g1 = [4.5 * cm] + [col_ancho] * len(periodos)
    g1_table = Table(g1_rows, colWidths=col_widths_g1, repeatRows=1)
    g1_table.setStyle(TableStyle(_make_th_style()))

    story.append(Paragraph('Sección 1 — Indicadores Globales de Planta', st_sec))
    story.append(g1_table)
    story.append(Spacer(1, 0.3 * cm))

    # Gráficas de la sección 1: G1+G1b en una fila, G2+G3 en otra
    _cr1 = _charts_row(['g1', 'g1b'], doc.width)
    if _cr1:
        story.append(_cr1)
        story.append(Spacer(1, 0.2 * cm))
    _cr2 = _charts_row(['g2', 'g3'], doc.width)
    if _cr2:
        story.append(_cr2)
        story.append(Spacer(1, 0.2 * cm))

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 2 — Por Línea / Máquina
    # Una sub-tabla por cada KPI; página nueva al inicio
    # ─────────────────────────────────────────────────────────────────────────
    por_grupo = datos.get('por_grupo', {})
    grupos    = por_grupo.get('grupos', [])
    agrup     = por_grupo.get('agrupado_por', 'linea')
    titulo_col = 'Línea' if agrup == 'linea' else 'Máquina'

    if grupos:
        story.append(PageBreak())
        story.append(Paragraph(f'Sección 2 — Indicadores por {titulo_col}', st_sec))

        TEND_MAP = {'mejora': '↑ Mejora', 'estable': '→ Estable', 'deterioro': '↓ Deterioro'}
        GRUPO_KPIS_PDF = [
            ('MTBF (h)',           'mtbf_h',   'tendencia_mtbf',   '{:.2f}'),
            ('MTTR (h)',           'mttr_h',   'tendencia_mttr',   '{:.2f}'),
            ('Disponibilidad (%)', 'disp_pct', 'tendencia_disp',   '{:.2f}'),
            ('Nº Paros',           'n_paros',  None,               '{:.0f}'),
            ('Horas de Paro',      'h_paros',  None,               '{:.2f}'),
            ('Tasa Fallos λ (f/h)','lambda',   'tendencia_lambda', '{:.5f}'),
        ]

        # Ancho columnas: grupo | periodos… | tendencia (opcional)
        c_grupo = 4.0 * cm
        c_tend  = 2.2 * cm
        n_p = max(len(periodos), 1)

        for kpi_label, kpi_campo, tend_key, fmt in GRUPO_KPIS_PDF:
            has_tend = tend_key is not None
            c_p = (doc.width - c_grupo - (c_tend if has_tend else 0)) / n_p
            header = [titulo_col] + periodos + (['Tendencia'] if has_tend else [])
            rows = [header]
            for g in grupos:
                fila = [g['nombre']]
                for p in g['periodos']:
                    v = p.get(kpi_campo)
                    fila.append(fmt.format(v) if v is not None else '—')
                if has_tend:
                    fila.append(TEND_MAP.get(g.get(tend_key, ''), '—'))
                rows.append(fila)

            cw = [c_grupo] + [c_p] * len(periodos) + ([c_tend] if has_tend else [])
            tbl = Table(rows, colWidths=cw, repeatRows=1)
            tbl.setStyle(TableStyle(_make_th_style()))

            story.append(KeepTogether([
                Paragraph(kpi_label, st_sub),
                tbl,
            ]))
            story.append(Spacer(1, 0.25 * cm))

        # Gráficas de la sección 2: G4 (disponibilidad) + G5 (nº paros)
        _cr_g45 = _charts_row(['g4', 'g5'], doc.width)
        if _cr_g45:
            story.append(Spacer(1, 0.1 * cm))
            story.append(_cr_g45)
            story.append(Spacer(1, 0.2 * cm))

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 3 — Top 10 Equipos
    # ─────────────────────────────────────────────────────────────────────────
    top10 = datos.get('top10', [])
    if top10:
        story.append(PageBreak())
        t10_rows = [['Equipo (Línea — Máquina)', 'Nº Paros', '% del Total', '% Acum.', 'Horas Paro']]
        for t in top10:
            equipo = (f"{t.get('linea','—')} — {t.get('maquina','')}"
                      if t.get('maquina') else t.get('linea', '—'))
            t10_rows.append([
                equipo,
                f"{t.get('n_paros', 0):.0f}",
                f"{t.get('pct_total', 0):.1f} %",
                f"{t.get('pct_acumulado', 0):.1f} %",
                f"{t.get('h_paros', 0):.2f} h",
            ])
        t10_table = Table(t10_rows,
                          colWidths=[8 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm],
                          repeatRows=1)
        t10_table.setStyle(TableStyle(_make_th_style()))

        story.append(KeepTogether([
            Paragraph('Sección 3 — Top 10 Equipos con más Paros', st_sec),
            t10_table,
        ]))
        story.append(Spacer(1, 0.3 * cm))

        # Gráfica G6 — Pareto de paros (ancho completo)
        _cr_g6 = _charts_row(['g6'], doc.width)
        if _cr_g6:
            story.append(_cr_g6)
            story.append(Spacer(1, 0.2 * cm))

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN 4 — Benchmarking Clase Mundial
    # ─────────────────────────────────────────────────────────────────────────
    bench_rows = datos.get('benchmarking', {}).get('rows', [])
    if bench_rows:
        bh = ['Indicador', 'Valor Actual', 'Referencia Clase Mundial', 'Gap', 'Estado']
        b_rows = [bh]
        ESTADO_MAP = {'ok': 'OK', 'mejorable': 'Mejorable', 'critico': 'Critico', 'nd': 'N/D'}
        for r in bench_rows:
            v = r.get('valor')
            b_rows.append([
                r.get('indicador', ''),
                f"{v:.3f} {r.get('unidad','')}" if v is not None else '—',
                r.get('referencia', ''),
                r.get('gap', ''),
                ESTADO_MAP.get(r.get('estado', ''), '—'),
            ])
        b_style = _make_th_style()
        for i, r in enumerate(bench_rows, start=1):
            est = r.get('estado')
            col = VERDE if est == 'ok' else NARANJA if est == 'mejorable' else ROJO if est == 'critico' else GRIS_TXT
            b_style += [
                ('TEXTCOLOR', (4, i), (4, i), col),
                ('FONTNAME',  (4, i), (4, i), 'Helvetica-Bold'),
            ]
        b_table = Table(b_rows,
                        colWidths=[5 * cm, 3.5 * cm, 4.5 * cm, 3 * cm, 3 * cm],
                        repeatRows=1)
        b_table.setStyle(TableStyle(b_style))

        story.append(KeepTogether([
            Paragraph('Sección 4 — Benchmarking Clase Mundial', st_sec),
            b_table,
        ]))

    doc.build(story, onFirstPage=_page_deco, onLaterPages=_page_deco)
    buf.seek(0)
    return buf

