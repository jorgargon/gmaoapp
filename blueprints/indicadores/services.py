"""
Lógica de negocio para el módulo de Indicadores y Reportes del GMAO.
Adaptado al modelo de datos real del proyecto (models.py).
"""
from datetime import datetime, date, timedelta
from sqlalchemy import func, or_, and_

from models import (
    db, OrdenTrabajo, ConsumoRecambio, RegistroTiempo,
    MovimientoStock, Recambio, Tecnico,
    GamaMantenimiento, AsignacionGama, TareaGama, RecambioGama,
    Maquina, Elemento, Linea, Zona, Planta, Empresa,
    ConfiguracionGeneral
)


# =============================================================================
# HELPERS INTERNOS
# =============================================================================

def _get_equipo_info(equipo_tipo, equipo_id):
    """Devuelve (codigo, nombre) para cualquier nivel jerárquico de activo."""
    modelos = {
        'empresa': Empresa, 'planta': Planta, 'zona': Zona,
        'linea': Linea, 'maquina': Maquina, 'elemento': Elemento
    }
    modelo = modelos.get(equipo_tipo)
    if not modelo or not equipo_id:
        return '', 'Sin asignar'
    obj = modelo.query.get(equipo_id)
    if not obj:
        return '', 'Desconocido'
    return getattr(obj, 'codigo', ''), getattr(obj, 'nombre', '')


def _build_tecnicos_dict():
    """
    Construye un dict {nombre_completo: coste_hora} para lookup rápido.
    Incluye claves con nombre completo y solo nombre.
    """
    result = {}
    for t in Tecnico.query.filter_by(activo=True).all():
        full = f"{t.nombre} {t.apellidos}".strip() if t.apellidos else t.nombre
        ch = t.costeHora or 0.0
        result[full] = ch
        result[t.nombre] = ch
    return result


def _coste_hora_defecto():
    """Coste/hora por defecto desde ConfiguracionGeneral o 0."""
    val = ConfiguracionGeneral.obtener('coste_hora_defecto', '0')
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _horas_y_coste_mo(orden, tecnicos_dict, coste_defecto):
    """
    Calcula horas de intervención y coste de mano de obra de una OT.
    Usa RegistroTiempo (entradas cerradas). Si no hay registros, usa tiempoReal.
    """
    horas = 0.0
    coste = 0.0
    registros_cerrados = [rt for rt in orden.registrosTiempo if rt.fin is not None]

    if registros_cerrados:
        for rt in registros_cerrados:
            h = (rt.fin - rt.inicio).total_seconds() / 3600
            horas += h
            nombre = (rt.tecnico or '').strip()
            ch = tecnicos_dict.get(nombre, coste_defecto)
            coste += h * ch
    else:
        # Fallback: usar tiempoReal y coste_defecto
        horas = orden.tiempoReal or 0.0
        coste = horas * coste_defecto

    return horas, coste


def _coste_recambios_orden(orden):
    """Suma coste de todos los recambios consumidos en una OT."""
    return sum((c.cantidad or 0) * (c.precioUnitario or 0) for c in orden.consumos)


def _parse_fecha(s):
    """Parsea 'YYYY-MM-DD' a date. Devuelve None si es inválida o vacía."""
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _frecuencia_a_delta(tipo, valor):
    """Convierte (tipo, valor) de frecuencia a timedelta."""
    valor = valor or 1
    if tipo == 'dias':
        return timedelta(days=valor)
    elif tipo == 'semanas':
        return timedelta(weeks=valor)
    elif tipo == 'meses':
        return timedelta(days=valor * 30)
    return timedelta(days=valor)


def _fmt(val, decimales=1):
    """Redondea un valor KPI. Devuelve None si es None."""
    if val is None:
        return None
    return round(val, decimales)


# =============================================================================
# JERARQUÍA DE ACTIVOS
# =============================================================================

def get_hijos_jerarquia(nivel, parent_id=None):
    """
    Devuelve lista de {id, codigo, nombre} para el nivel hijo del dado.
      nivel='root'    → empresas
      nivel='empresa' → plantas de esa empresa
      nivel='planta'  → zonas de esa planta
      nivel='zona'    → líneas de esa zona
      nivel='linea'   → máquinas de esa línea
      nivel='maquina' → elementos de esa máquina
    """
    _MAP = {
        'root':    (Empresa, None),
        'empresa': (Planta,  'empresaId'),
        'planta':  (Zona,    'plantaId'),
        'zona':    (Linea,   'zonaId'),
        'linea':   (Maquina, 'lineaId'),
        'maquina': (Elemento,'maquinaId'),
    }
    if nivel not in _MAP:
        return []

    modelo, fk_attr = _MAP[nivel]
    q = modelo.query
    if fk_attr and parent_id:
        q = q.filter(getattr(modelo, fk_attr) == int(parent_id))
    elif fk_attr:
        return []   # necesita parent_id

    items = q.order_by(modelo.nombre).all()
    return [{'id': i.id, 'codigo': getattr(i, 'codigo', ''), 'nombre': i.nombre} for i in items]


def _get_pares_bajo_nodo(nivel, nivel_id):
    """
    Devuelve lista de (equipoTipo, equipoId) para el nodo dado y todos sus descendientes.
    Recorre la jerarquía de forma recursiva: empresa→planta→zona→linea→maquina→elemento.
    """
    nivel_id = int(nivel_id)
    result = [(nivel, nivel_id)]

    if nivel == 'elemento':
        return result
    if nivel == 'maquina':
        for e in Elemento.query.filter_by(maquinaId=nivel_id).all():
            result.append(('elemento', e.id))
        return result
    if nivel == 'linea':
        for m in Maquina.query.filter_by(lineaId=nivel_id).all():
            result.extend(_get_pares_bajo_nodo('maquina', m.id))
        return result
    if nivel == 'zona':
        for l in Linea.query.filter_by(zonaId=nivel_id).all():
            result.extend(_get_pares_bajo_nodo('linea', l.id))
        return result
    if nivel == 'planta':
        for z in Zona.query.filter_by(plantaId=nivel_id).all():
            result.extend(_get_pares_bajo_nodo('zona', z.id))
        return result
    if nivel == 'empresa':
        for p in Planta.query.filter_by(empresaId=nivel_id).all():
            result.extend(_get_pares_bajo_nodo('planta', p.id))
        return result
    return result


# =============================================================================
# INFORME DE ÓRDENES DE TRABAJO
# =============================================================================

def get_informe_ordenes(fecha_inicio, fecha_fin, tipo=None, estado=None, equipo_id=None):
    """
    Devuelve (rows, totales) con datos de OT para el período indicado.
    rows: lista de dicts con todos los campos del informe.
    totales: dict con sumas de horas y costes.
    """
    tecnicos_dict = _build_tecnicos_dict()
    coste_defecto = _coste_hora_defecto()

    q = OrdenTrabajo.query

    if fecha_inicio:
        fi = datetime.combine(fecha_inicio, datetime.min.time())
        q = q.filter(OrdenTrabajo.fechaCreacion >= fi)
    if fecha_fin:
        ff = datetime.combine(fecha_fin, datetime.max.time())
        q = q.filter(OrdenTrabajo.fechaCreacion <= ff)
    if tipo:
        q = q.filter(OrdenTrabajo.tipo == tipo)
    if estado:
        q = q.filter(OrdenTrabajo.estado == estado)
    if equipo_id:
        q = q.filter(OrdenTrabajo.equipoId == int(equipo_id))

    ordenes = q.order_by(OrdenTrabajo.fechaCreacion.desc()).all()

    rows = []
    totales = {
        'horas_intervencion': 0.0,
        'horas_paro': 0.0,
        'coste_recambios': 0.0,
        'coste_talleres': 0.0,
        'coste_total': 0.0,
    }

    for o in ordenes:
        eq_codigo, eq_nombre = _get_equipo_info(o.equipoTipo, o.equipoId)
        horas, coste_mo = _horas_y_coste_mo(o, tecnicos_dict, coste_defecto)
        coste_rec = _coste_recambios_orden(o)
        coste_ext = o.costeTallerExterno or 0.0
        coste_total = coste_mo + coste_rec + coste_ext
        horas_paro = o.tiempoParada or 0.0

        totales['horas_intervencion'] += horas
        totales['horas_paro'] += horas_paro
        totales['coste_recambios'] += coste_rec
        totales['coste_talleres'] += coste_ext
        totales['coste_total'] += coste_total

        rows.append({
            'numero': o.numero,
            'titulo': o.titulo or '',
            'fecha_creacion': o.fechaCreacion.strftime('%d/%m/%Y') if o.fechaCreacion else '',
            'fecha_inicio': o.fechaInicio.strftime('%d/%m/%Y') if o.fechaInicio else '',
            'fecha_fin': o.fechaFin.strftime('%d/%m/%Y') if o.fechaFin else '',
            'equipo_codigo': eq_codigo,
            'equipo_nombre': eq_nombre,
            'tipo': o.tipo,
            'prioridad': o.prioridad,
            'estado': o.estado,
            'descripcion_averia': o.descripcionProblema or '',
            'trabajos_realizados': o.descripcionSolucion or '',
            'tecnico_asignado': o.tecnicoAsignado or '',
            'horas_intervencion': round(horas, 2),
            'horas_paro': round(horas_paro, 2),
            'coste_recambios': round(coste_rec, 2),
            'coste_talleres': round(coste_ext, 2),
            'coste_total': round(coste_total, 2),
        })

    # Redondear totales
    for k in totales:
        totales[k] = round(totales[k], 2)

    return rows, totales


def exportar_ordenes_excel(rows, totales):
    """Genera BytesIO con el Excel del informe de OT."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes de Trabajo'

    cabeceras = [
        'Nº Orden', 'Título', 'Fecha Solicitud', 'Fecha Inicio', 'Fecha Fin',
        'Equipo Código', 'Equipo', 'Tipo', 'Prioridad', 'Estado',
        'Descripción Avería', 'Trabajos Realizados', 'Técnico',
        'Horas Intervención', 'Horas Paro',
        'Coste Recambios (€)', 'Coste Talleres (€)', 'Coste Total (€)'
    ]
    campos = [
        'numero', 'titulo', 'fecha_creacion', 'fecha_inicio', 'fecha_fin',
        'equipo_codigo', 'equipo_nombre', 'tipo', 'prioridad', 'estado',
        'descripcion_averia', 'trabajos_realizados', 'tecnico_asignado',
        'horas_intervencion', 'horas_paro',
        'coste_recambios', 'coste_talleres', 'coste_total'
    ]

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    ws.row_dimensions[1].height = 30
    for col_idx, cab in enumerate(cabeceras, 1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, campo in enumerate(campos, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(campo, ''))

    # Fila de totales
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value='TOTALES').font = Font(bold=True)
    totales_cols = {14: 'horas_intervencion', 15: 'horas_paro',
                    16: 'coste_recambios', 17: 'coste_talleres', 18: 'coste_total'}
    for col, key in totales_cols.items():
        cell = ws.cell(row=total_row, column=col, value=totales.get(key, 0))
        cell.font = Font(bold=True)
        cell.fill = total_fill

    # Autoajustar anchos
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =============================================================================
# INFORME DE PREVENTIVOS PLANIFICADOS
# =============================================================================

def get_informe_preventivos(fecha_desde, fecha_hasta, equipo_id=None):
    """
    Devuelve (rows, resumen) con preventivos planificados en el rango.
    Incluye OTs ya creadas y futuras pendientes de crear (PENDIENTE).
    """
    rows = []

    asig_q = AsignacionGama.query.filter_by(activo=True)
    if equipo_id:
        asig_q = asig_q.filter(AsignacionGama.equipoId == int(equipo_id))

    asignaciones = asig_q.all()

    for asig in asignaciones:
        gama = asig.gama
        if not gama or not gama.activo:
            continue

        eq_codigo, eq_nombre = _get_equipo_info(asig.equipoTipo, asig.equipoId)

        # Tareas resumidas (primeras 3)
        tareas_str = '; '.join(t.descripcion[:40] for t in gama.tareas[:3])
        if len(gama.tareas) > 3:
            tareas_str += f' (+{len(gama.tareas) - 3} más)'

        # Recambios necesarios
        rec_items = [rg for rg in gama.recambios if rg.recambio]
        rec_str = '; '.join(
            f"{rg.recambio.nombre} x{rg.cantidad}"
            for rg in rec_items[:3]
        )
        if len(rec_items) > 3:
            rec_str += f' (+{len(rec_items) - 3} más)'

        # Horas previstas (tiempoEstimado en minutos)
        horas_previstas = (gama.tiempoEstimado or 0) / 60

        frecuencia_str = f"Cada {asig.frecuenciaValor} {asig.frecuenciaTipo}"
        delta = _frecuencia_a_delta(asig.frecuenciaTipo, asig.frecuenciaValor)

        # Generar fechas en el rango
        ultima = asig.ultimaEjecucion or fecha_desde
        fecha_actual = ultima
        fechas_en_rango = set()

        # Iterar hasta fecha_hasta (máx 500 iteraciones para evitar bucle infinito)
        for _ in range(500):
            fecha_actual = fecha_actual + delta
            if fecha_actual > fecha_hasta:
                break
            if fecha_actual >= fecha_desde:
                fechas_en_rango.add(fecha_actual)

        # Incluir proximaEjecucion si cae en el rango
        if asig.proximaEjecucion and fecha_desde <= asig.proximaEjecucion <= fecha_hasta:
            fechas_en_rango.add(asig.proximaEjecucion)

        for fecha_plan in sorted(fechas_en_rango):
            # Buscar OT existente para esta gama y equipo cerca de la fecha (±5 días)
            fecha_min = datetime.combine(fecha_plan - timedelta(days=5), datetime.min.time())
            fecha_max = datetime.combine(fecha_plan + timedelta(days=5), datetime.max.time())

            ot = OrdenTrabajo.query.filter(
                OrdenTrabajo.gamaId == gama.id,
                OrdenTrabajo.equipoId == asig.equipoId,
                OrdenTrabajo.equipoTipo == asig.equipoTipo,
                OrdenTrabajo.fechaProgramada >= fecha_min,
                OrdenTrabajo.fechaProgramada <= fecha_max,
            ).first()

            rows.append({
                'numero': ot.numero if ot else 'PENDIENTE',
                'fecha_planificada': fecha_plan.strftime('%d/%m/%Y'),
                'fecha_planificada_iso': fecha_plan.isoformat(),
                'equipo_codigo': eq_codigo,
                'equipo_nombre': eq_nombre,
                'gama_codigo': gama.codigo,
                'gama_nombre': gama.nombre,
                'frecuencia': frecuencia_str,
                'horas_previstas': round(horas_previstas, 2),
                'tareas': tareas_str,
                'recambios_necesarios': rec_str,
                'ultimo_preventivo': (
                    asig.ultimaEjecucion.strftime('%d/%m/%Y')
                    if asig.ultimaEjecucion else 'Nunca'
                ),
                'estado': ot.estado if ot else 'No creada',
                'es_pendiente': ot is None,
            })

    rows.sort(key=lambda x: x['fecha_planificada_iso'])

    resumen = {
        'total_planificados': len(rows),
        'total_horas_previstas': round(sum(r['horas_previstas'] for r in rows), 2),
        'pendientes': sum(1 for r in rows if r['es_pendiente']),
        'con_ot': sum(1 for r in rows if not r['es_pendiente']),
    }
    return rows, resumen


def exportar_preventivos_excel(rows):
    """Genera BytesIO con el Excel del informe de preventivos."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Preventivos Planificados'

    cabeceras = [
        'Nº Orden', 'Fecha Planificada', 'Equipo Código', 'Equipo',
        'Gama Código', 'Gama', 'Frecuencia', 'Horas Previstas',
        'Tareas', 'Recambios Necesarios', 'Último Preventivo', 'Estado'
    ]
    campos = [
        'numero', 'fecha_planificada', 'equipo_codigo', 'equipo_nombre',
        'gama_codigo', 'gama_nombre', 'frecuencia', 'horas_previstas',
        'tareas', 'recambios_necesarios', 'ultimo_preventivo', 'estado'
    ]

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    pending_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    ws.row_dimensions[1].height = 25
    for col_idx, cab in enumerate(cabeceras, 1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in enumerate(rows, 2):
        for col_idx, campo in enumerate(campos, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(campo, ''))
            if row.get('es_pendiente'):
                cell.fill = pending_fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =============================================================================
# INFORME DE MOVIMIENTOS DE STOCK
# =============================================================================

def get_informe_movimientos(fecha_inicio, fecha_fin, tipo=None, recambio_id=None):
    """
    Devuelve (rows, totales) con movimientos de stock en el período.
    """
    q = MovimientoStock.query

    if fecha_inicio:
        fi = datetime.combine(fecha_inicio, datetime.min.time())
        q = q.filter(MovimientoStock.fecha >= fi)
    if fecha_fin:
        ff = datetime.combine(fecha_fin, datetime.max.time())
        q = q.filter(MovimientoStock.fecha <= ff)
    if tipo:
        q = q.filter(MovimientoStock.tipo == tipo)
    if recambio_id:
        q = q.filter(MovimientoStock.recambioId == int(recambio_id))

    movs = q.order_by(MovimientoStock.fecha.desc()).all()

    rows = []
    totales = {
        'entradas_uds': 0,
        'salidas_uds': 0,
        'coste_entradas': 0.0,
        'coste_salidas': 0.0,
    }

    for m in movs:
        rec = m.recambio
        precio = rec.precioUnitario if rec else 0.0
        cant = m.cantidad or 0
        coste_total_mov = abs(cant) * precio

        tipo_display = {
            'entrada': 'ENTRADA',
            'salida': 'SALIDA',
            'ajuste': 'AJUSTE',
        }.get(m.tipo, (m.tipo or '').upper())

        if m.tipo == 'entrada':
            totales['entradas_uds'] += cant
            totales['coste_entradas'] += coste_total_mov
        elif m.tipo in ('salida', 'ajuste'):
            totales['salidas_uds'] += abs(cant)
            totales['coste_salidas'] += coste_total_mov

        # Número de OT desde documentoRef si es consumo
        orden_ref = ''
        if m.subTipo == 'consumo_ot' and m.documentoRef:
            orden_ref = m.documentoRef

        rows.append({
            'fecha': m.fecha.strftime('%d/%m/%Y %H:%M') if m.fecha else '',
            'tipo': tipo_display,
            'subtipo': m.subTipo or '',
            'recambio_codigo': rec.codigo if rec else '',
            'recambio_nombre': rec.nombre if rec else '',
            'cantidad': cant,
            'precio_unitario': round(precio, 2),
            'coste_total': round(coste_total_mov, 2),
            'orden_trabajo': orden_ref,
            'albaran': m.documentoRef if m.tipo == 'entrada' else '',
            'usuario': m.usuario or '',
            'motivo': m.motivo or '',
            'stock_anterior': m.stockAnterior if m.stockAnterior is not None else '',
            'stock_posterior': m.stockPosterior if m.stockPosterior is not None else '',
        })

    for k in ('coste_entradas', 'coste_salidas'):
        totales[k] = round(totales[k], 2)

    return rows, totales


def exportar_movimientos_excel(rows, totales):
    """Genera BytesIO con el Excel del informe de movimientos."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimientos de Stock'

    cabeceras = [
        'Fecha', 'Tipo', 'Subtipo', 'Código Recambio', 'Recambio',
        'Cantidad', 'Precio Unit. (€)', 'Coste Total (€)',
        'Orden Trabajo', 'Albarán', 'Usuario', 'Motivo',
        'Stock Anterior', 'Stock Posterior'
    ]
    campos = [
        'fecha', 'tipo', 'subtipo', 'recambio_codigo', 'recambio_nombre',
        'cantidad', 'precio_unitario', 'coste_total',
        'orden_trabajo', 'albaran', 'usuario', 'motivo',
        'stock_anterior', 'stock_posterior'
    ]

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    entrada_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    salida_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

    ws.row_dimensions[1].height = 25
    for col_idx, cab in enumerate(cabeceras, 1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in enumerate(rows, 2):
        row_fill = None
        if row.get('tipo') == 'ENTRADA':
            row_fill = entrada_fill
        elif row.get('tipo') in ('SALIDA', 'AJUSTE'):
            row_fill = salida_fill

        for col_idx, campo in enumerate(campos, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(campo, ''))
            if row_fill:
                cell.fill = row_fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =============================================================================
# INDICADORES KPI (EN 15341)
# =============================================================================

def calcular_indicadores(fecha_inicio, fecha_fin, nivel=None, nivel_id=None):
    """
    Calcula todos los KPI para el período y alcance jerárquico dados.
    nivel puede ser: empresa, planta, zona, linea, maquina, elemento (o None = toda la instalación).
    Devuelve dict agrupado: economicos, tecnicos, organizativos, resumen.
    """
    tecnicos_dict = _build_tecnicos_dict()
    coste_defecto = _coste_hora_defecto()

    fi = datetime.combine(fecha_inicio, datetime.min.time())
    ff = datetime.combine(fecha_fin, datetime.max.time())

    q = OrdenTrabajo.query.filter(
        OrdenTrabajo.fechaCreacion >= fi,
        OrdenTrabajo.fechaCreacion <= ff,
    )
    if nivel and nivel_id:
        pares = _get_pares_bajo_nodo(nivel, nivel_id)
        # Agrupar por tipo para construir OR eficiente
        by_tipo = {}
        for t, i in pares:
            by_tipo.setdefault(t, []).append(i)
        conditions = [
            and_(OrdenTrabajo.equipoTipo == t, OrdenTrabajo.equipoId.in_(ids))
            for t, ids in by_tipo.items()
        ]
        if conditions:
            q = q.filter(or_(*conditions))

    ordenes = q.all()

    # ── Clasificar OTs ────────────────────────────────────────────────────────
    correctivas = [o for o in ordenes if o.tipo == 'correctivo']
    preventivas = [o for o in ordenes if o.tipo == 'preventivo']

    # ── Acumular horas y costes ───────────────────────────────────────────────
    coste_total = 0.0
    coste_correctivo = 0.0
    coste_preventivo = 0.0
    coste_recambios_total = 0.0
    horas_total = 0.0
    horas_correctivo = 0.0
    horas_preventivo = 0.0
    horas_paro_total = 0.0

    for o in ordenes:
        horas, coste_mo = _horas_y_coste_mo(o, tecnicos_dict, coste_defecto)
        coste_rec = _coste_recambios_orden(o)
        coste_ext = o.costeTallerExterno or 0.0
        coste_ot = coste_mo + coste_rec + coste_ext
        hp = o.tiempoParada or 0.0

        coste_total += coste_ot
        coste_recambios_total += coste_rec
        horas_total += horas
        horas_paro_total += hp

        if o.tipo == 'correctivo':
            coste_correctivo += coste_ot
            horas_correctivo += horas
        elif o.tipo == 'preventivo':
            coste_preventivo += coste_ot
            horas_preventivo += horas

    # ── Parámetros temporales ─────────────────────────────────────────────────
    delta_dias = (fecha_fin - fecha_inicio).days + 1
    horas_periodo = delta_dias * 24  # Asume operación 24/7

    # ── Valor de stock de recambios ───────────────────────────────────────────
    valor_stock = db.session.query(
        func.sum(Recambio.stockActual * Recambio.precioUnitario)
    ).scalar() or 0.0

    # ── Indicadores Técnicos ──────────────────────────────────────────────────

    # T1 - Disponibilidad operacional
    tiempo_func = max(horas_periodo - horas_paro_total, 0.0)
    t1 = (tiempo_func / horas_periodo * 100) if horas_periodo > 0 else None

    # T3 - MTBF
    n_fallos = len(correctivas)
    t3 = tiempo_func / n_fallos if n_fallos > 0 else None

    # T4 - MTTR
    correctivas_cerradas = [o for o in correctivas if o.estado == 'cerrada']
    n_cerradas = len(correctivas_cerradas)
    horas_reparacion = sum(
        _horas_y_coste_mo(o, tecnicos_dict, coste_defecto)[0]
        for o in correctivas_cerradas
    )
    t4 = horas_reparacion / n_cerradas if n_cerradas > 0 else None

    # T8 - % horas preventivo
    t8 = (horas_preventivo / horas_total * 100) if horas_total > 0 else None

    # T9 - % horas correctivo
    t9 = (horas_correctivo / horas_total * 100) if horas_total > 0 else None

    # T12 - Cumplimiento plan preventivo
    prev_planificadas = [
        o for o in preventivas
        if o.fechaProgramada and fi <= o.fechaProgramada <= ff
    ]
    prev_completadas = [o for o in prev_planificadas if o.estado == 'cerrada']
    t12 = (
        len(prev_completadas) / len(prev_planificadas) * 100
        if prev_planificadas else None
    )

    # ── Indicadores Económicos ────────────────────────────────────────────────

    # E6 - % coste correctivo
    e6 = (coste_correctivo / coste_total * 100) if coste_total > 0 else None

    # E7 - % coste preventivo
    e7 = (coste_preventivo / coste_total * 100) if coste_total > 0 else None

    # E13 - % materiales (recambios) sobre coste total
    e13 = (coste_recambios_total / coste_total * 100) if coste_total > 0 else None

    # ── Indicadores Organizativos ─────────────────────────────────────────────

    # O10 - % OTs planificadas (preventivo) vs total
    o10 = (len(preventivas) / len(ordenes) * 100) if ordenes else None

    # O12 - Cumplimiento de programa (cerradas antes de su fecha programada)
    con_prog = [o for o in ordenes if o.fechaProgramada and fi <= o.fechaProgramada <= ff]
    completadas_en_tiempo = [
        o for o in con_prog
        if o.estado == 'cerrada' and o.fechaFin and o.fechaFin <= o.fechaProgramada
    ]
    o12 = (
        len(completadas_en_tiempo) / len(con_prog) * 100
        if con_prog else None
    )

    # O15 - Utilización del personal
    n_tecnicos = Tecnico.query.filter_by(activo=True).count()
    if n_tecnicos > 0:
        horas_disponibles = n_tecnicos * delta_dias * 8  # 8h/día de jornada
        o15 = (horas_total / horas_disponibles * 100) if horas_disponibles > 0 else None
    else:
        o15 = None

    # O18 - Calidad de datos en GMAO
    def _ot_completa(o):
        return (
            o.fechaInicio is not None and
            o.fechaFin is not None and
            (o.tiempoReal or 0) > 0 and
            bool(o.descripcionSolucion and o.descripcionSolucion.strip()) and
            len(o.registrosTiempo) > 0
        )

    ots_completas = sum(1 for o in ordenes if _ot_completa(o))
    o18 = (ots_completas / len(ordenes) * 100) if ordenes else None

    # ── Construir respuesta ───────────────────────────────────────────────────
    return {
        'economicos': {
            'E1': {
                'valor': None, 'unidad': '%', 'nombre': 'Coste Mtto / RAV (E1)',
                'nd': True, 'nota': 'Requiere campo "valor_reposicion" en equipos'
            },
            'E6': {'valor': _fmt(e6), 'unidad': '%', 'nombre': '% Coste correctivo (E6)'},
            'E7': {'valor': _fmt(e7), 'unidad': '%', 'nombre': '% Coste preventivo (E7)'},
            'E13': {'valor': _fmt(e13), 'unidad': '%', 'nombre': '% Coste materiales (E13)'},
            'E14': {
                'valor': None, 'unidad': '%', 'nombre': 'Stock / RAV (E14)',
                'nd': True, 'nota': 'Requiere campo "valor_reposicion" en equipos'
            },
        },
        'tecnicos': {
            'T1': {'valor': _fmt(t1), 'unidad': '%', 'nombre': 'Disponibilidad (T1)'},
            'T3': {'valor': _fmt(t3, 2), 'unidad': 'h', 'nombre': 'MTBF (T3)'},
            'T4': {'valor': _fmt(t4, 2), 'unidad': 'h', 'nombre': 'MTTR (T4)'},
            'T8': {'valor': _fmt(t8), 'unidad': '%', 'nombre': '% Horas preventivo (T8)'},
            'T9': {'valor': _fmt(t9), 'unidad': '%', 'nombre': '% Horas correctivo (T9)'},
            'T12': {'valor': _fmt(t12), 'unidad': '%', 'nombre': 'Cumpl. preventivo (T12)'},
        },
        'organizativos': {
            'O1': {
                'valor': None, 'unidad': '%', 'nombre': '% Personal interno (O1)',
                'nd': True, 'nota': 'Requiere campo "tipo_contrato" en técnicos'
            },
            'O10': {'valor': _fmt(o10), 'unidad': '%', 'nombre': '% OTs planificadas (O10)'},
            'O12': {'valor': _fmt(o12), 'unidad': '%', 'nombre': 'Cumpl. programa (O12)'},
            'O15': {'valor': _fmt(o15), 'unidad': '%', 'nombre': 'Utilización personal (O15)'},
            'O18': {'valor': _fmt(o18), 'unidad': '%', 'nombre': 'Calidad de datos (O18)'},
        },
        'resumen': {
            'total_ordenes': len(ordenes),
            'total_correctivas': len(correctivas),
            'total_preventivas': len(preventivas),
            'horas_total': round(horas_total, 1),
            'horas_paro': round(horas_paro_total, 1),
            'horas_periodo': horas_periodo,
            'coste_total': round(coste_total, 2),
            'coste_recambios': round(coste_recambios_total, 2),
            'coste_talleres': round(sum(o.costeTallerExterno or 0 for o in ordenes), 2),
            'valor_stock': round(valor_stock, 2),
            'delta_dias': delta_dias,
        }
    }
